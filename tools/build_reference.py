#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_reference.py

Data_list_217F.xlsx(사람이 관리하는 원본 엑셀, 입력지)를 읽어
분석 프로그램이 실행 시 참조할 data/*.json 파일들을 생성한다.

원칙
  - 엑셀 = 원본(single source of truth), JSON = 생성물.
  - 엑셀이 바뀌면 이 스크립트만 다시 실행해 JSON을 재생성한다.
  - 값은 엑셀에서 기계적으로 그대로 읽는다(추정/생성 금지).
  - 분석 프로그램은 실행 시 이 JSON만 로컬에서 읽으며 AI는 개입하지 않는다.

사용법
  python build_reference.py                       # 기본 경로로 실행
  python build_reference.py --excel <경로> --outdir data
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl 이 필요합니다:  pip install openpyxl")

# 유효성 목록에서 'Commercial 고정' 을 표시한 채우기 색(노란색)
FILL_COMMERCIAL_FIXED = "FFE699"

SHEET_CRITERIA = "217F 분석기준"
SHEET_VALIDITY = "유효성 목록"
SHEET_PSA = "PSA 입력 파라미터"

# 매핑맵 맨 앞 고정 3열(파라미터가 아닌 식별 컬럼). headers.json = 이 3열 + PSA 파라미터 union.
FIXED_HEADERS = ["품번\n(Part Number)", "Part Category", "Part Subcategory"]


# ----------------------------------------------------------------------
# 공통 유틸
# ----------------------------------------------------------------------
def _norm(v):
    return None if v is None else str(v).strip()


def find_cell(ws, exact=None, contains=None, within_col=None):
    """조건에 맞는 첫 셀 (row, col) 반환. 없으면 None."""
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if within_col is not None and c.column != within_col:
                continue
            s = str(c.value).strip()
            if exact is not None and s == exact:
                return c.row, c.column
            if contains is not None and contains in s:
                return c.row, c.column
    return None


def is_commercial_fixed(cell):
    """셀 채우기 색이 Commercial 고정 표시색인지."""
    try:
        rgb = cell.fill.fgColor.rgb
    except Exception:
        return False
    return rgb is not None and FILL_COMMERCIAL_FIXED in str(rgb)


def num(v):
    """숫자면 int/float 로, 아니면 원본 그대로."""
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else float(v)
    return v


def cell_value(ws, row, col):
    """병합된 셀이면 병합 범위의 왼쪽 위(앵커) 셀 값을 돌려줘요. openpyxl은 병합된 칸 중
    앵커가 아닌 나머지는 그냥 .cell()로 읽으면 항상 None을 주기 때문에, 서브카테고리 이름처럼
    여러 파라미터 컬럼에 걸쳐 병합된 셀은 이렇게 읽어야 해요."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return ws.cell(row=row, column=col).value


# ----------------------------------------------------------------------
# 1) 운용 온도별 품질등급  ->  quality_by_temp.json
# ----------------------------------------------------------------------
def build_quality_by_temp(ws):
    # 헤더의 '구분' 셀을 찾되, 오른쪽에 '0℃' 온도구간이 오는 표를 앵커로 잡는다.
    anchor = None
    for row in ws.iter_rows():
        for c in row:
            if _norm(c.value) == "구분":
                right = ws.cell(row=c.row, column=c.column + 1).value
                if right and "℃" in str(right):
                    anchor = (c.row, c.column)
                    break
        if anchor:
            break
    if not anchor:
        raise RuntimeError("'운용 온도별 품질등급' 표(구분/온도구간)를 찾지 못함")

    hr, c0 = anchor
    ranges = []
    col = c0 + 1
    while True:
        v = _norm(ws.cell(row=hr, column=col).value)
        if not v:
            break
        ranges.append(v)
        col += 1

    mapping = {}
    r = hr + 1
    while True:
        cat = _norm(ws.cell(row=r, column=c0).value)
        if not cat:
            break
        mapping[cat] = {
            rng: _norm(ws.cell(row=r, column=c0 + 1 + i).value)
            for i, rng in enumerate(ranges)
        }
        r += 1

    return {
        "source": f"{SHEET_CRITERIA} / 운용 온도별 품질등급",
        "note": "카테고리 x 동작온도범위 -> 적용 품질등급. (Inductor는 Coil/Transformer 구분)",
        "temperature_ranges": ranges,
        "mapping": mapping,
    }


# ----------------------------------------------------------------------
# 2) θJC (Junction-to-Case)  ->  thermal_resistance.json
# ----------------------------------------------------------------------
def build_thermal_resistance(ws):
    pkg = find_cell(ws, exact="Package Type")
    if not pkg:
        raise RuntimeError("θJC 표(Package Type)를 찾지 못함")
    hr, pkg_col = pkg
    cat_col = pkg_col - 1  # 'Category'
    tr_col = pkg_col + 1   # 'Thermal Resistance'

    by_cat = {}
    r = hr + 1
    while True:
        cat = _norm(ws.cell(row=r, column=cat_col).value)
        pk = _norm(ws.cell(row=r, column=pkg_col).value)
        tr = ws.cell(row=r, column=tr_col).value
        if not cat and not pk:
            break
        if cat and pk and tr is not None:
            by_cat.setdefault(cat, {})[pk] = num(tr)
        r += 1

    return {
        "source": f"{SHEET_CRITERIA} / Junction-to-Case Thermal Resistance",
        "unit": "°C/W",
        "note": "데이터시트에 θJC 없을 때 사용. TJ = TC + θJC x P. 스펙값과 표값 중 큰 쪽 사용, 없으면 기본 70(반도체).",
        "by_category": by_cat,
    }


# ----------------------------------------------------------------------
# 3)/4) 변환 매트릭스 (환경/온도)  ->  env_conversion.json / temp_conversion.json
# ----------------------------------------------------------------------
def build_matrix(ws, to_anchor_text, from_anchor_text, source_desc, key_name):
    to_a = find_cell(ws, exact=to_anchor_text)
    from_a = find_cell(ws, exact=from_anchor_text)
    if not to_a or not from_a:
        raise RuntimeError(f"매트릭스 앵커({to_anchor_text}/{from_anchor_text})를 찾지 못함")

    # To 라벨: to_anchor 바로 아래 행, 그 오른쪽 열부터
    to_label_row = to_a[0] + 1
    to_cols = []
    col = to_a[1] + 1
    while True:
        v = _norm(ws.cell(row=to_label_row, column=col).value)
        if v is None or v == "":
            break
        to_cols.append((str(v), col))
        col += 1

    # From 라벨: from_anchor 오른쪽 열, from_anchor 행부터 아래로
    from_label_col = from_a[1] + 1
    from_rows = []
    row = from_a[0]
    while True:
        v = _norm(ws.cell(row=row, column=from_label_col).value)
        if v is None or v == "":
            break
        from_rows.append((str(v), row))
        row += 1

    matrix = {}
    for f_label, f_row in from_rows:
        matrix[f_label] = {
            t_label: num(ws.cell(row=f_row, column=t_col).value)
            for t_label, t_col in to_cols
        }

    return {
        "source": source_desc,
        "note": "적용값 = 데이터시트값 x 환경변환팩터 x 온도변환팩터. matrix[from][to] 형식. 표에 없는 값은 반올림하여 가장 가까운 칸 사용.",
        key_name: [lbl for lbl, _ in from_rows],
        "matrix": matrix,
    }


# ----------------------------------------------------------------------
# 5) 서브카테고리별 허용 품질등급  ->  quality_allowed.json
# ----------------------------------------------------------------------
def build_quality_allowed(ws):
    # 카테고리 마커(2행)
    cat_marks = sorted(
        (c.column, str(c.value).strip())
        for c in ws[2] if c.value not in (None, "")
    )

    def cat_of(col):
        name = None
        for cc, nm in cat_marks:
            if cc <= col:
                name = nm
            else:
                break
        return name

    by_cat = {}
    n_fixed = 0
    for c in ws[6]:  # 파라미터명 행
        if _norm(c.value) != "Quality Level":
            continue
        col = c.column
        subname = _norm(ws.cell(row=4, column=col).value)
        if not subname:
            continue
        # 허용값: 7행부터 아래로
        allowed = []
        tokens = []
        r = 7
        while True:
            v = _norm(ws.cell(row=r, column=col).value)
            if v is None or v == "":
                break
            allowed.append(v)
            tokens.extend(t.strip() for t in v.split(",") if t.strip())
            r += 1
        fixed = is_commercial_fixed(ws.cell(row=4, column=col)) or \
                is_commercial_fixed(ws.cell(row=6, column=col))
        if fixed:
            n_fixed += 1
        by_cat.setdefault(cat_of(col), {})[subname] = {
            "allowed": allowed,
            "allowed_tokens": sorted(set(tokens)),
            "commercial_fixed": fixed,
        }

    return {
        "source": f"{SHEET_VALIDITY} / 서브카테고리별 허용 Quality Level",
        "note": ("commercial_fixed=true 는 온도 기반 등급이 이 서브카테고리에서 미허용이라 "
                 "Quality Level 을 Commercial 로 고정하는 항목(규칙2). "
                 f"엑셀에서 채우기색 {FILL_COMMERCIAL_FIXED} 로 표시됨."),
        "commercial_fixed_count": n_fixed,
        "by_category": by_cat,
    }


# ----------------------------------------------------------------------
# 5-1) 서브카테고리별 허용 Package Type  ->  package_type_allowed.json
# ----------------------------------------------------------------------
def build_package_type_allowed(ws):
    """Package Type은 실제 물리적 패키지명(예: 'UTQFN-12')이 아니라, 217F가 정한 표준
    패키지 분류(예: 'Nonhermetic: DIPs, PGA, SMT')예요. build_quality_allowed와 완전히
    같은 방식(4행=서브카테고리명, 6행=파라미터명, 7행부터 아래로 허용값)으로 읽어요."""
    cat_marks = sorted(
        (c.column, str(c.value).strip())
        for c in ws[2] if c.value not in (None, "")
    )

    def cat_of(col):
        name = None
        for cc, nm in cat_marks:
            if cc <= col:
                name = nm
            else:
                break
        return name

    by_cat = {}
    for c in ws[6]:
        if _norm(c.value) != "Package Type":
            continue
        col = c.column
        subname = _norm(cell_value(ws, 4, col))
        if not subname:
            continue
        allowed = []
        r = 7
        while True:
            v = _norm(ws.cell(row=r, column=col).value)
            if v is None or v == "":
                break
            allowed.append(v)
            r += 1
        by_cat.setdefault(cat_of(col), {})[subname] = allowed

    return {
        "source": f"{SHEET_VALIDITY} / 서브카테고리별 허용 Package Type",
        "note": ("Package Type 파라미터가 이 서브카테고리에서 가질 수 있는 217F 표준 분류값 "
                 "목록(엑셀 드롭다운과 동일). 실제 물리적 패키지명이 아니에요 - 물리적 패키지를"
                 "이 목록 중 하나로 매핑하는 규칙은 ai/reference.py의 classify_package_hermeticity 참고."),
        "by_category": by_cat,
    }


# ----------------------------------------------------------------------
# 5-2) 서브카테고리별 허용값 전체 (모든 파라미터)  ->  valid_values.json
# ----------------------------------------------------------------------
def build_valid_values(ws):
    """'유효성 목록' 시트에 허용값 목록이 있는 모든 파라미터를 통째로 읽어요(Quality Level/
    Package Type뿐 아니라 Type/Units/Junction-/Technology Type/Construction Type 등 전부).

    CLAUDE.md 0번 원칙(2026-08-27 확정): 이 시트에 허용값이 정의된 파라미터는, 그 값 중에서만
    골라야 해요. 이 JSON은 ai/reference.py의 get_allowed_values()/is_allowed_value()가 읽어서,
    분석 파이프라인 마지막에 모든 필드값을 이 목록과 대조하는 안전망으로 써요 - 목록에 없는
    값은 자동으로 비우고(사람 확인 필요), 이 시트에 애초에 없는 파라미터(전압/전류 등 연속값)는
    이 규칙 대상이 아니니 그대로 둬요.

    build_quality_allowed/build_package_type_allowed와 같은 표 구조(2행=카테고리 마커,
    4행=서브카테고리명(병합 셀), 6행=파라미터명, 7행부터 아래로 허용값)를 파라미터 이름 상관없이
    전부 훑어요."""
    cat_marks = sorted(
        (c.column, str(c.value).strip())
        for c in ws[2] if c.value not in (None, "")
    )

    def cat_of(col):
        name = None
        for cc, nm in cat_marks:
            if cc <= col:
                name = nm
            else:
                break
        return name

    by_cat = {}
    param_count = 0
    for c in ws[6]:
        pname = _norm(c.value)
        if not pname:
            continue
        col = c.column
        subname = _norm(cell_value(ws, 4, col))
        if not subname:
            continue
        allowed = []
        r = 7
        while True:
            v = _norm(ws.cell(row=r, column=col).value)
            if v is None or v == "":
                break
            allowed.append(v)
            r += 1
        if not allowed:
            continue
        cat = cat_of(col)
        by_cat.setdefault(cat, {}).setdefault(subname, {})[pname] = allowed
        param_count += 1

    return {
        "source": f"{SHEET_VALIDITY} (전체 파라미터)",
        "note": ("카테고리/서브카테고리/파라미터별 허용값 전체 목록. 이 목록에 있는 파라미터의 "
                 "값은 반드시 이 중에서만 골라야 함(CLAUDE.md 0번 원칙, 2026-08-27). 여기 없는 "
                 "파라미터(연속 측정값 등)는 이 규칙 대상이 아님."),
        "param_column_count": param_count,
        "by_category": by_cat,
    }


# ----------------------------------------------------------------------
# 6) PSA 파라미터 목록  ->  subcat_params.json
# ----------------------------------------------------------------------
def build_subcat_params(ws):
    # 헤더행에서 Category/Subcategory 열 위치 찾기
    hdr = find_cell(ws, exact="Subcategory")
    if not hdr:
        raise RuntimeError("PSA 시트에서 'Subcategory' 헤더를 찾지 못함")
    hr, sub_col = hdr
    cat_col = sub_col - 1  # Category
    param_start = sub_col + 1  # 파라미터 나열 시작 열

    # 파라미터 열은 전체 서브카테고리가 공유하는 고정 슬롯(헤더행에 SD25, SD26, ... 처럼
    # 이름이 매겨져 있음)이라, 어떤 행에서는 슬롯 사이사이가 비어요(그 서브카테고리엔 해당
    # 안 되는 파라미터라서) - 그래서 "빈 칸을 만나면 끝"이 아니라, 헤더행에 실제로 슬롯이
    # 있는 마지막 열까지는 전부 훑고 빈 칸은 그냥 건너뛰어야 해요(2026-09-04 수정 - PSA
    # 입력 파라미터 시트는 그대로 두고 읽는 방식만 고침, 사용자 확정). 예전 방식(첫 빈 칸에서
    # 멈춤)은 Transistor 행에서 Operating Voltage 다음 칸이 비어 있다는 이유로 그 뒤의 Rated
    # Voltage/Voltage Ratio/Power Rating/Operating Power/Thermal Resistance/Junction-/
    # Temperature Rise/Junction Temp Override를 전부 놓치고 있었음(실제 확인).
    param_end = param_start
    col = param_start
    while _norm(ws.cell(row=hr, column=col).value):
        param_end = col
        col += 1

    items = []
    for r in range(hr + 1, ws.max_row + 1):
        cat = _norm(ws.cell(row=r, column=cat_col).value)
        sub = _norm(ws.cell(row=r, column=sub_col).value)
        if not cat or not sub:
            continue
        params = [
            v
            for col in range(param_start, param_end + 1)
            if (v := _norm(ws.cell(row=r, column=col).value))
        ]
        items.append({"category": cat, "subcategory": sub, "params": params})

    return items


# ----------------------------------------------------------------------
# 7) 매핑맵 컬럼 헤더  ->  headers.json
# ----------------------------------------------------------------------
def build_headers(subcat_items):
    """subcat_params 의 파라미터를 처음 등장 순서대로 모아 매핑맵 헤더를 만든다.

    subcat_params 와 같은 원본(PSA 시트)에서 파생되므로, 모든 파라미터가 headers 에
    반드시 존재하도록 구조적으로 보장된다(이원화/철자 불일치 방지). 고정 3열이 맨 앞에 온다.
    """
    order = []
    seen = set()
    for e in subcat_items:
        for p in e["params"]:
            if p not in seen:
                seen.add(p)
                order.append(p)
    return FIXED_HEADERS + order


# ----------------------------------------------------------------------
# 실행
# ----------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default="Data_list_217F.xlsx")
    ap.add_argument("--outdir", default="data")
    args = ap.parse_args()

    xl = Path(args.excel)
    if not xl.exists():
        sys.exit(f"엑셀 파일 없음: {xl}")
    out = Path(args.outdir)
    out.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xl)  # 스타일(채우기색) 접근 위해 기본 로드
    ws_c = wb[SHEET_CRITERIA]
    ws_v = wb[SHEET_VALIDITY]
    ws_p = wb[SHEET_PSA]

    subcat_items = build_subcat_params(ws_p)
    outputs = {
        "quality_by_temp.json": build_quality_by_temp(ws_c),
        "thermal_resistance.json": build_thermal_resistance(ws_c),
        "env_conversion.json": build_matrix(
            ws_c, "To Environment", "From Environment",
            f"{SHEET_CRITERIA} / 환경 변환 팩터", "environments"),
        "temp_conversion.json": build_matrix(
            ws_c, "To Temperature", "From Temperature",
            f"{SHEET_CRITERIA} / 온도 변환 팩터", "temperatures"),
        "quality_allowed.json": build_quality_allowed(ws_v),
        "package_type_allowed.json": build_package_type_allowed(ws_v),
        "valid_values.json": build_valid_values(ws_v),
        "subcat_params.json": subcat_items,
        # headers 는 subcat_params 에서 파생 — 모든 파라미터가 headers 에 포함됨을 구조적으로 보장.
        "headers.json": build_headers(subcat_items),
    }

    for name, data in outputs.items():
        p = out / name
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  생성: {p}")

    _verify(outputs)


def _verify(outputs):
    print("\n=== 검증 (원본 엑셀과 대조용 샘플) ===")
    q = outputs["quality_by_temp.json"]["mapping"]
    print("[quality_by_temp] 카테고리 수:", len(q),
          "| Capacitor(-55~125):", q.get("Capacitor", {}).get("-55℃ ~ 125℃"))

    tr = outputs["thermal_resistance.json"]["by_category"]
    ic = tr.get("Integrated Circuit", {})
    sc = tr.get("Semiconductor", {})
    print("[thermal_resistance] IC 패키지:", len(ic), "개 (Can=", ic.get("Can"),
          ") | 반도체:", len(sc), "개 (TO-220=", sc.get("TO-220"), ")")

    tc = outputs["temp_conversion.json"]["matrix"]
    print("[temp_conversion] 대각선 모두 1:",
          all(tc[str(t)][str(t)] == 1 for t in [10, 20, 30, 40, 50, 60, 70]),
          "| 30->40:", tc["30"]["40"])

    ec = outputs["env_conversion.json"]["matrix"]
    print("[env_conversion] GB->NS:", ec["GB"]["NS"], "| 환경 수:",
          len(outputs["env_conversion.json"]["environments"]))

    qa = outputs["quality_allowed.json"]
    print("[quality_allowed] 서브카테고리 수:",
          sum(len(v) for v in qa["by_category"].values()),
          "| Commercial 고정:", qa["commercial_fixed_count"], "개")

    pa = outputs["package_type_allowed.json"]
    print("[package_type_allowed] 서브카테고리 수:",
          sum(len(v) for v in pa["by_category"].values()),
          "| Linear:", pa["by_category"].get("Integrated Circuit", {}).get("Linear"))

    vv = outputs["valid_values.json"]
    print("[valid_values] 파라미터 칸 수:", vv["param_column_count"],
          "| Linear Package Type:",
          vv["by_category"].get("Integrated Circuit", {}).get("Linear", {}).get("Package Type"))

    sp = outputs["subcat_params.json"]
    print("[subcat_params] 항목 수:", len(sp),
          "| 예:", sp[0]["category"], "/", sp[0]["subcategory"],
          "->", len(sp[0]["params"]), "파라미터")

    hd = outputs["headers.json"]
    all_params = {p for e in sp for p in e["params"]}
    missing = all_params - set(hd)
    print("[headers] 총 컬럼:", len(hd), "(고정3 + 파라미터", len(hd) - len(FIXED_HEADERS),
          ") | subcat_params 모든 파라미터 포함:", not missing)


if __name__ == "__main__":
    main()
