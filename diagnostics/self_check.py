# 프로그램이 스스로 이상 여부를 점검하는 파일이에요 (부가기능 2번: 자가 검진 코드).
# 승인된 4가지를 각각 함수로 나눠뒀어요.
#  a) check_workbook_integrity - 매핑맵 색상/서식이 CLAUDE.md 규칙에서 벗어나지 않았는지
#     (과거 LibreOffice 재저장으로 색상이 통째로 사라진 사고의 재발을 조기에 잡기 위함)
#  b) check_connectivity - Mouser API 키/인터넷 연결, 다운로드 폴더 쓰기 권한
#  c) check_field_anomalies - 추출된 값이 음수/단위누락/상식 밖 범위인지
#  d) check_output_file - 방금 만든 출력 엑셀/PDF가 다시 정상적으로 열리는지

import re
import tempfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pymupdf as fitz  # "import fitz"는 옛 이름이라 매번 경고가 떠서 새 이름으로 불러옴
import requests

from ai.field_extractor import FIELD_SYNONYMS
from ai.prompt import load_headers, load_subcat_params
from excel.mapping_colors import WHITE_HEX, normalize_argb
from excel.mapping_template_builder import (
    DATA_START_ROW,
    FIXED_COLUMN_COUNT,
    HEADER_ROW,
    MAPPING_SHEET_NAME,
    expected_hex_for,
)
from utils.config import DOWNLOAD_DIR, ENV_SEARCH_PATHS, MOUSER_API_KEY


@dataclass
class Issue:
    level: str  # "오류" | "경고"
    message: str


def check_workbook_integrity(path: Path, sample_rows: int = 30) -> list[Issue]:
    if not path.exists():
        return [Issue("오류", f"파일이 없습니다: {path}")]
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        return [Issue("오류", f"엑셀 파일을 열 수 없습니다: {e}")]

    issues: list[Issue] = []
    try:
        if MAPPING_SHEET_NAME not in wb.sheetnames:
            return [Issue("오류", f"'{MAPPING_SHEET_NAME}' 시트가 없습니다.")]
        ws = wb[MAPPING_SHEET_NAME]

        expected_headers = load_headers()
        headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
        if headers[: len(expected_headers)] != expected_headers:
            issues.append(Issue("오류", "매핑맵 헤더가 data/headers.json과 일치하지 않습니다."))

        # 실제로 값을 적어 넣은 흰 칸은 "칠 없음"으로 저장된 경우가 있어(엑셀의 기본 배경이라 굳이
        # 색을 지정하지 않음), 흰색은 검사 대상에서 빼요. 회색/보라/분홍처럼 사람이 의도적으로 칠한
        # "특별한" 칸만 확인하면, 과거처럼 색이 통째로 사라지는 사고를 오탐 없이 잡을 수 있어요.
        subcat_lookup = {(e["category"], e["subcategory"]): set(e["params"]) for e in load_subcat_params()}

        checked = 0
        mismatches = 0
        last_row = min(ws.max_row, DATA_START_ROW + sample_rows - 1)
        for r in range(DATA_START_ROW, last_row + 1):
            category = ws.cell(row=r, column=2).value
            subcategory = ws.cell(row=r, column=3).value
            applicable = subcat_lookup.get((category, subcategory))
            if applicable is None:
                continue  # 이 행의 대분류/소분류를 모르면 기대색을 알 수 없으니 판단을 보류해요.

            for col, name in enumerate(expected_headers[FIXED_COLUMN_COUNT:], start=FIXED_COLUMN_COUNT + 1):
                expected = expected_hex_for(name, applicable)
                if expected == WHITE_HEX:
                    continue
                cell = ws.cell(row=r, column=col)
                actual = normalize_argb(cell.fill.fgColor.rgb) if cell.fill and cell.fill.patternType else None
                checked += 1
                if actual != expected:
                    mismatches += 1
        if checked and mismatches:
            issues.append(
                Issue(
                    "경고",
                    f"매핑맵 표본 {checked}칸(회/보라/분홍이어야 할 칸) 중 {mismatches}칸이 기대한 색과 다릅니다 "
                    "(서식 손상 의심).",
                )
            )
    finally:
        wb.close()
    return issues


def check_connectivity(download_dir: Path | None = None) -> list[Issue]:
    issues: list[Issue] = []

    if not MOUSER_API_KEY:
        checked = " / ".join(str(p) for p in ENV_SEARCH_PATHS)
        issues.append(Issue("오류", f"MOUSER_API_KEY가 .env에 설정되어 있지 않습니다. (확인한 위치: {checked})"))
    else:
        try:
            resp = requests.post(
                "https://api.mouser.com/api/v1/search/keyword",
                params={"apiKey": MOUSER_API_KEY},
                json={"SearchByKeywordRequest": {"keyword": "TEST", "records": 1}},
                timeout=8,
            )
            if resp.status_code in (401, 403):
                issues.append(Issue("오류", "Mouser API 키가 거부되었습니다 (인증 오류)."))
            elif resp.status_code >= 500:
                issues.append(Issue("경고", f"Mouser API가 서버 오류를 응답했습니다 (HTTP {resp.status_code})."))
        except requests.RequestException as e:
            issues.append(Issue("오류", f"Mouser API/인터넷 연결 확인 실패: {e}"))

    target_dir = download_dir or DOWNLOAD_DIR
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(dir=target_dir, delete=True):
            pass
    except OSError as e:
        issues.append(Issue("오류", f"다운로드 폴더에 쓸 수 없습니다({target_dir}): {e}"))

    return issues


_NUMERIC_RE = re.compile(r"[-+]?\d[\d,]*\.?\d*")
_UNIT_RE = re.compile(r"[a-zA-Zµμ°%Ω]")

# 필드별 "이 값을 넘으면 이상해 보인다" 상식적 상한선이에요 (데이터시트에 흔한 표기 기준).
_MAX_REASONABLE = {
    "Operating Voltage": 1000,
    "Rated Voltage": 2000,
    "Operating Current": 100,
    "Rated Current": 200,
    "Power Rating": 2000,
    "Operating Power": 100000,  # mW 단위로 저장하는 관례(CLAUDE.md 7번)를 감안한 상한
    "Thermal Resistance": 1000,
    "Frequency": 1e12,
    "Max Junction Temp": 300,
    "Operating Temperature": 300,
}


def check_field_anomalies(field_values: dict[str, str | None]) -> list[Issue]:
    issues: list[Issue] = []
    for field_name, raw_value in field_values.items():
        if not raw_value or field_name not in FIELD_SYNONYMS:
            continue  # 글자값(카테고리형) 필드는 숫자 이상치 검사 대상이 아니에요.

        match = _NUMERIC_RE.search(raw_value)
        if not match:
            continue
        try:
            number = float(match.group().replace(",", ""))
        except ValueError:
            continue

        if number < 0:
            issues.append(Issue("경고", f"{field_name} 값이 음수입니다: {raw_value!r}"))
        if not _UNIT_RE.search(raw_value):
            issues.append(Issue("경고", f"{field_name} 값에 단위가 없어 보입니다: {raw_value!r}"))

        limit = _MAX_REASONABLE.get(field_name)
        if limit and abs(number) > limit:
            issues.append(
                Issue("경고", f"{field_name} 값이 상식적 범위를 벗어난 것 같습니다: {raw_value!r} (기준 {limit})")
            )

    return issues


def check_output_file(xlsx_path: Path, pdf_paths: list[Path]) -> list[Issue]:
    issues = check_workbook_integrity(xlsx_path)

    for pdf_path in pdf_paths:
        try:
            doc = fitz.open(pdf_path)
            try:
                if doc.page_count == 0:
                    issues.append(Issue("오류", f"{pdf_path.name}: 페이지가 0개입니다."))
            finally:
                doc.close()
        except Exception as e:
            issues.append(Issue("오류", f"{pdf_path.name}을 다시 열 수 없습니다: {e}"))

    return issues
