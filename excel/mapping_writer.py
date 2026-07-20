# 분석 결과를 Windchill 매핑맵 시트에 실제로 써넣는 파일이에요.
#
# 지켜야 할 규칙 (CLAUDE.md / '사용가이드라인' 시트에서 그대로 가져옴):
# - 흰 칸(직접 입력)만 채운다. 회색(N/A)/보라(PTC 자동결정)/분홍(데이터시트로 알 수 없음) 칸은 절대 안 건드림.
# - 같은 서브카테고리에 부품이 이미 있으면, 그 행의 서식(색칠)을 그대로 복제해서 바로 아래에 새 행을 추가.
# - LibreOffice로 다시 저장하는 것 절대 금지 — openpyxl로 열고 값만 바꾼 뒤 그대로 저장.

from copy import copy

import openpyxl

from ai.prompt import get_fields_for_subcategory

MAPPING_SHEET_NAME = "매핑맵"
HEADER_ROW = 2
DATA_START_ROW = 3
COL_PART_NUMBER = 1
COL_CATEGORY = 2
COL_SUBCATEGORY = 3


def _find_rows_for_subcategory(ws, category: str, subcategory: str) -> list[int]:
    # 이 대분류/소분류와 일치하는 행 번호를 순서대로 찾아요.
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_CATEGORY).value == category and ws.cell(row=r, column=COL_SUBCATEGORY).value == subcategory:
            rows.append(r)
    return rows


def _clone_row_style(ws, src_row: int, dst_row: int, max_col: int):
    # 값은 복제하지 않고, 색깔/글꼴/테두리 같은 "서식"만 그대로 복제해요.
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def find_or_create_row(ws, category: str, subcategory: str) -> int:
    """이 부품의 값을 넣을 행 번호를 찾거나(비어있는 템플릿 행), 없으면 새로 만들어요.

    이미 다른 부품이 그 서브카테고리를 다 쓰고 있으면, 마지막 행의 서식을 그대로 복제해서
    바로 아래에 새 행을 추가해요 (부품이 1개뿐이던 서브카테고리도 똑같이 처리돼요).
    """
    rows = _find_rows_for_subcategory(ws, category, subcategory)
    if not rows:
        raise ValueError(f"'{category}' / '{subcategory}' 서브카테고리를 매핑맵에서 찾을 수 없습니다.")

    for r in rows:
        if ws.cell(row=r, column=COL_PART_NUMBER).value in (None, ""):
            return r  # 아직 아무 부품도 안 들어간 빈 템플릿 행

    last_row = rows[-1]
    new_row = last_row + 1
    ws.insert_rows(new_row)
    _clone_row_style(ws, last_row, new_row, ws.max_column)
    # insert_rows는 서식이 없는 빈 행만 만들어주므로, 대분류/소분류 값도 다시 채워줘야 해요.
    ws.cell(row=new_row, column=COL_CATEGORY, value=category)
    ws.cell(row=new_row, column=COL_SUBCATEGORY, value=subcategory)
    return new_row


def write_part_to_mapping(ws, row: int, part_number: str, category: str, subcategory: str, field_values: dict):
    """이 행의 흰 칸(입력 가능한 칸)에만 값을 채워요. 다른 색 칸은 절대 안 건드려요."""
    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}

    ws.cell(row=row, column=COL_PART_NUMBER, value=part_number)

    fillable = set(get_fields_for_subcategory(category, subcategory))
    for field, value in field_values.items():
        if field not in fillable or value is None:
            continue  # 이 서브카테고리엔 없는 칸이거나, 못 찾은 값이면 건드리지 않아요.
        col = header_to_col.get(field)
        if col:
            ws.cell(row=row, column=col, value=value)


class MappingMapWriter:
    """매핑맵 엑셀 파일을 한 번 열어두고, 부품을 하나씩 써넣은 뒤 저장하는 역할을 해요."""

    def __init__(self, path: str):
        self.path = path
        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb[MAPPING_SHEET_NAME]

    def write_part(self, part_number: str, category: str, subcategory: str, field_values: dict) -> int:
        row = find_or_create_row(self.ws, category, subcategory)
        write_part_to_mapping(self.ws, row, part_number, category, subcategory, field_values)
        return row

    def save(self):
        self.wb.save(self.path)

    def close(self):
        self.wb.close()
