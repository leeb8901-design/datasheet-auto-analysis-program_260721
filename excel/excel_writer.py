# 처리 결과(다운로드 상태, 분석 상태 등)를 원본 엑셀 파일에 다시 써넣는 파일이에요.

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from excel import psa_writer
from utils.config import COL_DATASHEET_LINK, PART_NUMBER_KEYWORDS, WORKSHEET_RESULT_COLUMN_MAP

# 엑셀에서 링크처럼 보이도록 파란색 밑줄 글씨체를 만들어둬요.
_HYPERLINK_FONT = Font(color="0563C1", underline="single")

_HEADER_SCAN_ROWS = 15  # 헤더가 1행이 아닐 수 있어(작업지는 2행) 앞쪽 이 범위에서 헤더를 찾아요.


class ExcelResultWriter:
    """엑셀 파일을 한 번 열어두고, 행마다 결과를 채워 넣은 뒤 저장하는 역할을 해요.

    분석 결과는 config의 WORKSHEET_RESULT_COLUMN_MAP에 따라 '작업지'의 정해진 컬럼에 써넣어요
    (다운로드 상태 -> '상태', 데이터시트 링크 -> '데이터시트 다운로드 링크',
    저장 경로 -> '분석된 데이터시트 링크'). 매핑에 없는 결과값은 작업지에 쓰지 않아요.
    """

    def __init__(self, path: str, sheet_name: str):
        self.path = path
        self.sheet_name = sheet_name
        # read_only가 아니라 진짜로 "쓰기"가 가능한 모드로 열어요.
        # .xlsm(매크로 포함) 파일은 keep_vba=True로 열어야 해요. 이거 없이 열고 저장하면
        # openpyxl이 VBA 프로젝트(datasheet_helper.bas 매크로)를 통째로 지워버려서, 나중에
        # Import 양식에서 다운로드 도우미 매크로를 실행하려 해도 매크로 자체가 사라져 있는
        # 문제가 있었어요.
        keep_vba = str(path).lower().endswith(".xlsm")
        self.wb = openpyxl.load_workbook(path, keep_vba=keep_vba)
        self.ws = self.wb[sheet_name]
        self.header_row = self._detect_header_row()
        self.column_map = self._ensure_result_columns()
        self._datasheet_key = None  # PSA 시트의 '데이터시트' 범례 색 키(첫 사용 시 계산)

    def _detect_header_row(self) -> int:
        # 헤더가 1행이 아닐 수 있어요(작업지는 2행). '품번'이 있는 줄을 헤더로 찾아요(reader와 동일 기준).
        for r in range(1, _HEADER_SCAN_ROWS + 1):
            for c in range(1, self.ws.max_column + 1):
                v = self.ws.cell(row=r, column=c).value
                if v and any(k in str(v).strip().lower() for k in PART_NUMBER_KEYWORDS):
                    return r
        return 1

    def _ensure_result_columns(self) -> dict[str, int]:
        # 헤더 줄에 이미 있는 컬럼 이름과 번호를 읽어와요.
        headers: dict[str, int] = {}
        max_col = self.ws.max_column
        for col in range(1, max_col + 1):
            value = self.ws.cell(row=self.header_row, column=col).value
            if value:
                headers[str(value).strip()] = col

        # 매핑 대상 컬럼(상태/데이터시트 다운로드 링크/분석된 데이터시트 링크)이 없으면 헤더 줄에 만들어요.
        next_col = max_col + 1
        for target in WORKSHEET_RESULT_COLUMN_MAP.values():
            if target not in headers:
                self.ws.cell(row=self.header_row, column=next_col, value=target)
                headers[target] = next_col
                next_col += 1

        self.wb.save(self.path)
        return headers

    def write_row(
        self,
        row_index: int,
        values: dict[str, str],
        link_path: Path | None = None,
        reference_url: str | None = None,
    ):
        # values 예: {"다운로드 상태": "성공 (Mouser)", "데이터시트 링크": "AD8030ARZ.pdf", ...}
        # 각 결과키를 WORKSHEET_RESULT_COLUMN_MAP으로 작업지 컬럼(상태 등)에 매핑해서 써요.
        # 매핑에 없는 결과(분석 상태/오류내용/미확인 항목)는 건너뛰어요.
        # link_path/reference_url을 주면 '데이터시트 다운로드 링크' 칸이 클릭 가능한 링크가 돼요.
        for key, value in values.items():
            target = WORKSHEET_RESULT_COLUMN_MAP.get(key)
            if target is None:
                continue
            col = self.column_map.get(target)
            if not col:
                continue
            cell = self.ws.cell(row=row_index, column=col, value=value)
            if key != COL_DATASHEET_LINK:
                continue
            if link_path is not None:
                # 일반 윈도우 경로(C:\...) 형태로 넣어야 엑셀에서 클릭했을 때 바로 열려요.
                # file:/// 형태(URI)로 넣으면 경로에 한글이 있을 때 인코딩이 깨져서 "파일을 열 수 없다"는
                # 오류가 나는 경우가 있어서, 그냥 실제 경로 문자열을 그대로 써요.
                cell.hyperlink = str(link_path.resolve())
                cell.font = _HYPERLINK_FONT
            elif reference_url:
                cell.hyperlink = reference_url
                cell.font = _HYPERLINK_FONT

    def write_part_params(self, category, subcategory, part_number, part_name, field_values):
        """분석된 부품의 파라미터를 같은 workbook의 'PSA 입력 파라미터' 시트에 써넣어요.
        해당 소분류 정의행 아래에 부품 행을 추가하고, '데이터시트' 색 파라미터만 값(없으면 노란색)을 채워요.
        PSA 시트가 없거나 소분류 정의행을 못 찾으면 조용히 넘어가요."""
        if not category or not subcategory:
            return
        if psa_writer.PSA_SHEET_NAME not in self.wb.sheetnames:
            return
        ws = self.wb[psa_writer.PSA_SHEET_NAME]
        if self._datasheet_key is None:
            self._datasheet_key = psa_writer.get_datasheet_color_key(ws)
        psa_writer.write_part(
            ws, category, subcategory, part_number, part_name, field_values, self._datasheet_key
        )

    def save(self):
        # 지금까지의 변경사항을 실제 파일에 저장해요. (Excel 자동 저장 요구사항)
        self.wb.save(self.path)

    def close(self):
        self.wb.close()
