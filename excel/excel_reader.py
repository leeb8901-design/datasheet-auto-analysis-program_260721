# 엑셀 파일을 열어서, 그 안에 적힌 부품번호/제조사를 읽어오는 파일이에요.

import openpyxl

from utils.config import (
    MANUFACTURER_KEYWORDS,
    PART_LIST_SHEET_NAME,
    PART_NAME_KEYWORDS,
    PART_NUMBER_KEYWORDS,
)

# 헤더(제목 줄)가 항상 1행에 있는 건 아니에요. 위에 빈 줄이나 제목/설명 줄이 먼저 오는 입력지도
# 있어서(예: Data_list_217F의 부품리스트는 헤더가 2행), 앞쪽 이 범위 안에서 '품번' 헤더가 있는
# 줄을 찾아 헤더 행으로 삼아요.
HEADER_SCAN_ROWS = 15


def get_sheet_names(path: str) -> list[str]:
    # 엑셀 파일 안에는 시트(탭)가 여러 개 있을 수 있어요. 그 이름들을 전부 알려줘요.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def get_headers(path: str, sheet_name: str) -> list[str]:
    # 고른 시트의 맨 첫 줄(제목 줄)에 뭐라고 적혀있는지 읽어와요.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(h).strip() if h is not None else "" for h in header_row]
    finally:
        wb.close()


def find_column(headers: list[str], keywords: list[str]) -> int | None:
    # 제목 줄에서 힌트 단어가 들어있는 칸을 찾아요. 못 찾으면 None을 돌려줘요.
    for i, h in enumerate(headers):
        low = h.lower()
        if any(k in low for k in keywords):
            return i
    return None


def guess_part_number_column(headers: list[str]) -> int:
    # find_column과 비슷하지만, 못 찾으면 0번째 칸을 기본값으로 써요 (수동 선택 팝업용).
    col = find_column(headers, PART_NUMBER_KEYWORDS)
    return col if col is not None else 0


def find_header_row(path: str, sheet_name: str) -> tuple[int, list[str]] | None:
    """앞쪽 여러 행을 훑어 '품번' 헤더가 있는 줄을 찾아 (행번호, 헤더리스트)를 돌려줘요.
    헤더가 1행이든 2행이든(위에 빈 줄/제목 줄이 있어도) 견고하게 찾기 위한 거예요. 못 찾으면 None."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True), start=1
        ):
            headers = [str(h).strip() if h is not None else "" for h in row]
            if find_column(headers, PART_NUMBER_KEYWORDS) is not None:
                return row_idx, headers
        return None
    finally:
        wb.close()


def find_input_columns(path: str, sheet_name: str) -> dict | None:
    """'부품번호'/'제조사' 컬럼이 실제 엑셀에서 몇 번째 컬럼(1-indexed, openpyxl 기준)인지
    찾아줘요. 프로그램 창에서 셀을 직접 고쳤을 때 원본 엑셀의 정확히 같은 칸에 되써넣기
    위해 필요해요. 헤더를 못 찾으면 None."""
    found = find_header_row(path, sheet_name)
    if found is None:
        return None
    header_row, headers = found
    part_col = find_column(headers, PART_NUMBER_KEYWORDS)
    if part_col is None:
        return None
    mfr_col = find_column(headers, MANUFACTURER_KEYWORDS)
    return {
        "header_row": header_row,
        "part_col": part_col + 1,  # openpyxl은 1-indexed
        "mfr_col": (mfr_col + 1) if mfr_col is not None else None,
    }


def read_part_list_sheet(path: str) -> list[dict] | None:
    """'부품리스트' 시트에서 '부품번호'와 '제조사' 컬럼을 찾아 읽어온다.
    이 시트가 없거나 부품번호 컬럼을 찾을 수 없으면 None을 반환한다.
    헤더 줄은 앞쪽 몇 행 안에서 자동으로 찾으므로, 헤더가 1행이 아니어도(예: 2행) 잘 읽는다.
    각 행은 {"row": 엑셀상의 실제 행 번호, "part_number": ..., "manufacturer": ... 또는 None} 형태다.
    같은 부품번호가 여러 번 나오면 처음 것만 남기고 건너뛴다(중복 제거).
    """
    sheets = get_sheet_names(path)
    if PART_LIST_SHEET_NAME not in sheets:
        return None

    found = find_header_row(path, PART_LIST_SHEET_NAME)
    if found is None:
        return None
    header_row, headers = found
    part_col = find_column(headers, PART_NUMBER_KEYWORDS)  # find_header_row가 보장하므로 None 아님
    mfr_col = find_column(headers, MANUFACTURER_KEYWORDS)
    name_col = find_column(headers, PART_NAME_KEYWORDS)

    # 데이터는 헤더 바로 다음 줄부터예요.
    return _read_rows(path, PART_LIST_SHEET_NAME, part_col, mfr_col, start_row=header_row + 1, name_col=name_col)


def read_custom_sheet(
    path: str, sheet_name: str, part_col: int, mfr_col: int | None, has_header: bool = True
) -> list[dict]:
    # 사용자가 팝업에서 직접 시트/컬럼을 고른 경우 쓰는 함수예요.
    return _read_rows(path, sheet_name, part_col, mfr_col, start_row=(2 if has_header else 1))


def _read_rows(path, sheet_name, part_col, mfr_col, start_row, name_col=None) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = []
        seen = set()  # 이미 넣은 부품번호를 기억해서 중복을 걸러내요.
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if part_col >= len(row):
                continue
            part_value = row[part_col]
            if part_value is None:
                continue
            part_text = str(part_value).strip()
            if not part_text:
                continue
            key = part_text.lower()
            if key in seen:
                continue  # 중복 품번은 건너뛰어요.
            seen.add(key)

            mfr_text = None
            if mfr_col is not None and mfr_col < len(row) and row[mfr_col] is not None:
                mfr_text = str(row[mfr_col]).strip() or None

            name_text = None
            if name_col is not None and name_col < len(row) and row[name_col] is not None:
                name_text = str(row[name_col]).strip() or None

            rows.append({"row": row_idx, "part_number": part_text, "manufacturer": mfr_text, "part_name": name_text})
        return rows
    finally:
        wb.close()
