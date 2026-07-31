# '매핑맵' 시트를 처음부터 새로 만드는 파일이에요.
#
# 마스터 파일(Windchill_217F_Mapping_Template.xlsx)은 이미 여러 부품이 채워진 상태라, 서브카테고리마다
# 남아있던 "빈 행"이 이미 소진된 경우가 있어요. 그래서 출력지를 만들 때마다 마스터를 복사하는 대신,
# 검증된 data/headers.json + data/subcat_params.json(그리고 색상 규칙)만으로 완전히 새로운 빈 매핑맵을
# 만들어요. 마스터 파일은 이 과정에서 전혀 열지 않으므로, 마스터의 서식이 깨질 위험이 없어요.

import openpyxl.worksheet.worksheet

from ai.prompt import PTC_AUTO_DETERMINED_FIELDS, UNKNOWN_FROM_DATASHEET_FIELDS, load_headers, load_subcat_params
from excel.mapping_colors import FILL_BY_HEX, GRAY_HEX, PINK_HEX, PURPLE_HEX, WHITE_HEX

MAPPING_SHEET_NAME = "매핑맵"
LEGEND_ROW = 1
HEADER_ROW = 2
DATA_START_ROW = 3
FIXED_COLUMN_COUNT = 3  # 품번 / Part Category / Part Subcategory (색칠 대상이 아님)

LEGEND_ENTRIES = [
    (WHITE_HEX, "직접 입력 - 데이터시트에서 값을 찾아 채움"),
    (GRAY_HEX, "N/A - 해당 서브카테고리엔 적용 안 됨 (값을 넣지 않음)"),
    (PURPLE_HEX, "PTC 자동결정 - Windchill이 내부적으로 계산 (값을 넣지 않음)"),
    (PINK_HEX, "데이터시트로 알 수 없음 - 별도 확인 전까지 항상 공란"),
]


def expected_hex_for(field_name: str, applicable_fields: set[str]) -> str:
    # CLAUDE.md 3번 표의 하드코딩 규칙: 적용 안 되면 회색, 적용되면 보라/분홍 세트에 있는지 확인 후
    # 아니면 흰색. mapping_template_builder(생성)와 diagnostics/self_check(검증)가 똑같이 이 함수를
    # 써서, "만들 때 규칙"과 "검사할 때 규칙"이 어긋날 일이 없게 해요.
    if field_name not in applicable_fields:
        return GRAY_HEX
    if field_name in PTC_AUTO_DETERMINED_FIELDS:
        return PURPLE_HEX
    if field_name in UNKNOWN_FROM_DATASHEET_FIELDS:
        return PINK_HEX
    return WHITE_HEX


def build_blank_mapping_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    """새 워크북에 서브카테고리마다 빈 행이 색칠된 '매핑맵' 시트를 만들어 돌려줘요(개수는 subcat_params 기준)."""
    ws = wb.create_sheet(MAPPING_SHEET_NAME)
    headers = load_headers()  # [품번, Part Category, Part Subcategory, PSA 파라미터들...]

    for col, (hex_code, text) in enumerate(LEGEND_ENTRIES, start=1):
        cell = ws.cell(row=LEGEND_ROW, column=col, value=text)
        cell.fill = FILL_BY_HEX[hex_code]

    for col, name in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col, value=name)

    row = DATA_START_ROW
    for entry in load_subcat_params():
        applicable = set(entry["params"])
        ws.cell(row=row, column=2, value=entry["category"])
        ws.cell(row=row, column=3, value=entry["subcategory"])
        for col, name in enumerate(headers[FIXED_COLUMN_COUNT:], start=FIXED_COLUMN_COUNT + 1):
            ws.cell(row=row, column=col).fill = FILL_BY_HEX[expected_hex_for(name, applicable)]
        row += 1

    return ws
