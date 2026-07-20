# 매핑맵 시트의 4색 규칙(CLAUDE.md 3번 표)을 한 곳에 모아둔 파일이에요.
# mapping_writer.py, mapping_template_builder.py, diagnostics/self_check.py가 전부 이 값을 가져다 써요
# (같은 HEX가 여러 파일에 따로 적혀있으면, 나중에 하나만 고치고 다른 곳을 빠뜨리는 사고가 나기 쉬워서예요).

from openpyxl.styles import PatternFill

WHITE_HEX = "FFFFFF"  # 직접 입력
GRAY_HEX = "D9D9D9"  # N/A (해당없음, 절대 값 넣지 않음)
PURPLE_HEX = "D9C6EC"  # PTC 자동결정 (Windchill이 내부적으로 계산, 절대 값 넣지 않음)
PINK_HEX = "F8CBAD"  # 데이터시트로 알 수 없음 (항상 공란)

ALL_RULE_HEX = {WHITE_HEX, GRAY_HEX, PURPLE_HEX, PINK_HEX}

WHITE_FILL = PatternFill(fgColor=WHITE_HEX, fill_type="solid")
GRAY_FILL = PatternFill(fgColor=GRAY_HEX, fill_type="solid")
PURPLE_FILL = PatternFill(fgColor=PURPLE_HEX, fill_type="solid")
PINK_FILL = PatternFill(fgColor=PINK_HEX, fill_type="solid")

FILL_BY_HEX = {WHITE_HEX: WHITE_FILL, GRAY_HEX: GRAY_FILL, PURPLE_HEX: PURPLE_FILL, PINK_HEX: PINK_FILL}


def normalize_argb(value: str | None) -> str | None:
    """openpyxl이 돌려주는 fill.fgColor.rgb는 보통 'FFFFFFFF' 같은 8자리(ARGB)예요.
    앞의 알파 채널 2자리를 떼고 뒤 6자리(RGB)만 비교할 수 있게 정리해줘요."""
    if not value or not isinstance(value, str):
        return None
    return value[-6:].upper()
