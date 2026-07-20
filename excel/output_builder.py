# '출력지'라는 새 엑셀 파일을 만드는 파일이에요. 실행할 때마다 새로 만들고, 마스터 템플릿
# (Windchill_217F_Mapping_Template.xlsx)은 읽기만 할 뿐 절대 저장하지 않아요.
#
# 시트 3개로 구성돼요.
#  1) 사용가이드라인 - 마스터 파일에서 값+서식을 그대로 복사만 해와요 (정적 설명 시트라 안전함, CLAUDE.md 참고).
#  2) 부품리스트 - 이번 배치의 입력지 내용 + 다운로드/분석 처리 결과.
#  3) 매핑맵 - mapping_template_builder로 새로 만든 빈 매핑맵에, 이번 배치 부품들의 분석값을 채움.

from copy import copy
from pathlib import Path

import openpyxl

from ai.analysis_state import PartAnalysis
from excel.mapping_template_builder import build_blank_mapping_sheet
from excel.mapping_writer import find_or_create_row, write_part_to_mapping
from utils.config import (
    COL_ANALYSIS_STATUS,
    COL_DATASHEET_LINK,
    COL_DOWNLOAD_STATUS,
    COL_ERROR_MESSAGE,
    COL_UNRESOLVED_FIELDS,
)

GUIDE_SHEET_NAME = "사용가이드라인"
PART_LIST_SHEET_NAME = "부품리스트"


def _copy_guide_sheet(master_path: Path, target_wb) -> None:
    # 마스터 파일을 열어서 값만 읽어와요. 이 함수 안에서 master_wb.save()는 절대 호출하지 않아요.
    if not master_path.exists():
        return
    master_wb = openpyxl.load_workbook(master_path)
    try:
        if GUIDE_SHEET_NAME not in master_wb.sheetnames:
            return
        src = master_wb[GUIDE_SHEET_NAME]
        dst = target_wb.create_sheet(GUIDE_SHEET_NAME)
        for row in src.iter_rows():
            for cell in row:
                new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                new_cell.fill = copy(cell.fill)
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
        for col_letter, dim in src.column_dimensions.items():
            dst.column_dimensions[col_letter].width = dim.width
        for idx, dim in src.row_dimensions.items():
            dst.row_dimensions[idx].height = dim.height
    finally:
        master_wb.close()


def _build_part_list_sheet(target_wb, rows: list[dict], table_rows: list[dict]) -> None:
    ws = target_wb.create_sheet(PART_LIST_SHEET_NAME)
    ws.append(
        ["No.", "품번", "제조사", COL_DOWNLOAD_STATUS, COL_ANALYSIS_STATUS, COL_DATASHEET_LINK, COL_ERROR_MESSAGE, COL_UNRESOLVED_FIELDS]
    )
    for i, row in enumerate(rows):
        t = table_rows[i]
        ws.append(
            [
                i + 1,
                row["part_number"],
                t.get("manufacturer", ""),
                t.get("download_status", ""),
                t.get("analysis_status", ""),
                t.get("filename", ""),
                t.get("error", ""),
                t.get("unresolved", ""),
            ]
        )


def build_output_workbook(
    rows: list[dict],
    table_rows: list[dict],
    analysis_by_row: list[PartAnalysis | None],
    master_template_path: Path,
    save_path: Path,
) -> Path:
    """rows / table_rows / analysis_by_row는 모두 같은 순서(인덱스)로 짝지어져 있어야 해요."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본으로 생기는 빈 시트("Sheet")는 지워요.

    _copy_guide_sheet(master_template_path, wb)
    _build_part_list_sheet(wb, rows, table_rows)
    mapping_ws = build_blank_mapping_sheet(wb)

    for row, analysis in zip(rows, analysis_by_row):
        if analysis is None or not analysis.category or not analysis.subcategory:
            continue
        try:
            target_row = find_or_create_row(mapping_ws, analysis.category, analysis.subcategory)
        except ValueError:
            continue  # 서브카테고리를 못 찾으면(오분류 등) 매핑맵엔 건너뛰고 부품리스트 기록만 남겨요.
        write_part_to_mapping(
            mapping_ws, target_row, row["part_number"], analysis.category, analysis.subcategory, analysis.fields
        )

    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    wb.close()
    return save_path
