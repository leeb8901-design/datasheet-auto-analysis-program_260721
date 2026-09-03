# 분석 결과를 작업지(Data_list_217F.xlsx)의 'PSA 입력 파라미터' 시트에 써넣는 파일이에요.
#
# 규칙 (사용자 확정, 2026-07-31):
# - 각 서브카테고리 '정의행'(SD25~ 슬롯에 파라미터 '이름'이 있는 행) 아래에 부품마다 한 행씩 추가해요.
# - 정의행 셀 색이 범례의 '관련자료(Datasheet) 등 제시 및 분석결과값 적용' 스와치 색(=L5 색)과
#   같은 파라미터에만 분석값을 채워요. 그 외(Relex 자동계산/설계자 검토/Default)는 앱이 안 건드려요.
# - 데이터시트 대상 파라미터인데 값을 확인 못했으면(공백) 노란색으로 표시해요.
# - 값을 채운 칸엔 색을 안 칠해요(2026-08-27부터) - 정의행과 같은 색을 칠하면 글자와 배경이 겹쳐
#   안 보인다는 피드백이 있었어요(예: Quality Level 칸 색과 "B-1" 글자 구분이 안 됨).

from openpyxl.styles import PatternFill

PSA_SHEET_NAME = "PSA 입력 파라미터"
HEADER_ROW = 9

# 헤더행(9) 기준 식별 컬럼
COL_PMYEONG = 6    # F 품명
COL_PARTNUM = 7    # G PartNumber
COL_CATEGORY = 9   # I Category
COL_SUBCAT = 10    # J Subcategory
SD_START = 11      # K (SD25) 부터 파라미터 슬롯

YELLOW_FILL = PatternFill("solid", fgColor="FFFFFF00")  # 확인 못한 데이터시트 항목 표시색(범례 노랑과 동일)


def color_key(cell):
    """셀 채우기 색을 비교 가능한 키로 바꿔요(테마/rgb/indexed 모두). 색 없으면 None."""
    fill = cell.fill
    if not fill or not fill.patternType:
        return None
    c = fill.fgColor
    if c is None:
        return None
    if c.type == "theme":
        return ("theme", c.theme, round(float(c.tint or 0.0), 2))
    if c.type == "rgb" and isinstance(c.rgb, str):
        return ("rgb", c.rgb)
    if c.type == "indexed":
        return ("indexed", c.indexed)
    return None


def get_datasheet_color_key(ws):
    """범례에서 '관련자료(Datasheet)' 줄을 찾아, 그 줄의 스와치(색칠된 셀) 색 키를 돌려줘요."""
    for r in range(1, HEADER_ROW):
        for c in range(1, 25):
            v = ws.cell(row=r, column=c).value
            if v and "Datasheet" in str(v):
                for cc in range(1, 25):
                    key = color_key(ws.cell(row=r, column=cc))
                    if key:
                        return key
    return None


def find_definition_row(ws, category, subcategory):
    """(category, subcategory) 정의행을 찾아요(Category+Subcategory가 채워진 행). 없으면 None."""
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        if (ws.cell(row=r, column=COL_CATEGORY).value == category
                and ws.cell(row=r, column=COL_SUBCAT).value == subcategory):
            return r
    return None


def _insert_index(ws, def_row):
    """정의행 아래, 기존 부품 행들 뒤(다음 정의행 전)의 삽입 위치를 돌려줘요.
    부품 행은 Category가 비어있고 PartNumber가 있는 행이에요."""
    insert_at = def_row + 1
    r = def_row + 1
    while r <= ws.max_row:
        if ws.cell(row=r, column=COL_CATEGORY).value:  # 다음 정의행 만남
            break
        if ws.cell(row=r, column=COL_PARTNUM).value:   # 기존 부품 행
            insert_at = r + 1
        r += 1
    return insert_at


def write_part(ws, category, subcategory, part_number, part_name, field_values, datasheet_key):
    """정의행 아래에 부품 행을 추가하고 데이터시트색 파라미터 값을 채워요.
    돌려주는 값: 부품 행 번호(성공) 또는 None(정의행 못 찾음)."""
    def_row = find_definition_row(ws, category, subcategory)
    if def_row is None:
        return None

    at = _insert_index(ws, def_row)
    ws.insert_rows(at, 1)

    # 식별열 (Category/Subcategory는 넣지 않아요 - 정의행과 구분하기 위해)
    if part_name:
        ws.cell(row=at, column=COL_PMYEONG, value=part_name)
    ws.cell(row=at, column=COL_PARTNUM, value=part_number)

    # SD 파라미터 슬롯: 정의행 셀 색이 데이터시트색인 것만 처리
    for c in range(SD_START, ws.max_column + 1):
        def_cell = ws.cell(row=def_row, column=c)
        pname = def_cell.value
        if not pname:
            continue
        if datasheet_key is None or color_key(def_cell) != datasheet_key:
            continue  # 데이터시트 대상 아님 -> 비워둠
        value = field_values.get(str(pname).strip())
        pcell = ws.cell(row=at, column=c)
        if value is not None and str(value).strip() != "":
            pcell.value = value
            # 채운 값 칸엔 색을 안 칠해요 - 정의행과 같은 색을 칠했더니 글자색/배경색이 겹쳐서
            # 값이 잘 안 보인다는 문제가 있었어요(사용자 피드백, 예: Quality Level 칸 색과 "B-1"
            # 글자가 구분이 안 됨). 값이 있다는 건 그 자체로 눈에 띄니 굳이 안 칠해도 됨.
        else:
            pcell.fill = YELLOW_FILL          # 확인 못함 -> 공백 + 노란색
    return at
