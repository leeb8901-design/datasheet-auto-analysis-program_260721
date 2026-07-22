import openpyxl

from excel.excel_writer import ExcelResultWriter
from utils.config import COL_DATASHEET_LINK, COL_DOWNLOAD_STATUS, COL_LANDING_PAGE


def _make_input_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "부품리스트"
    ws.append(["No.", "품번", "제조사"])
    ws.append([1, "AD8030ARZ", ""])
    wb.save(path)
    wb.close()


def test_write_row_sets_landing_page_hyperlink(tmp_path):
    path = tmp_path / "import.xlsx"
    _make_input_workbook(path)

    writer = ExcelResultWriter(str(path), "부품리스트")
    writer.write_row(
        2,
        {
            COL_DOWNLOAD_STATUS: "실패",
            COL_DATASHEET_LINK: "https://www.analog.com/blocked.pdf",
            COL_LANDING_PAGE: "https://www.analog.com/en/products/ad8030.html",
        },
        reference_url="https://www.analog.com/blocked.pdf",
        landing_url="https://www.analog.com/en/products/ad8030.html",
    )
    writer.save()
    writer.close()

    wb = openpyxl.load_workbook(path)
    ws = wb["부품리스트"]
    col = writer.column_map[COL_LANDING_PAGE]
    cell = ws.cell(row=2, column=col)
    assert cell.value == "https://www.analog.com/en/products/ad8030.html"
    assert cell.hyperlink.target == "https://www.analog.com/en/products/ad8030.html"
    wb.close()


def test_write_row_skips_landing_hyperlink_when_no_landing_url(tmp_path):
    path = tmp_path / "import.xlsx"
    _make_input_workbook(path)

    writer = ExcelResultWriter(str(path), "부품리스트")
    writer.write_row(2, {COL_LANDING_PAGE: ""}, landing_url=None)
    writer.save()
    writer.close()

    wb = openpyxl.load_workbook(path)
    ws = wb["부품리스트"]
    col = writer.column_map[COL_LANDING_PAGE]
    cell = ws.cell(row=2, column=col)
    assert cell.hyperlink is None
    wb.close()
