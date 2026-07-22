import openpyxl

import ui.main_window as main_window_module
from datasheet.downloader import DownloadResult
from ui.main_window import DatasheetWorker
from utils.config import COL_LANDING_PAGE, STATUS_FAILED


class _StubMouserClient:
    def __init__(self, *args, **kwargs):
        pass


def _make_input_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "부품리스트"
    ws.append(["No.", "품번", "제조사"])
    ws.append([1, "AD8030ARZ", ""])
    wb.save(path)
    wb.close()


def test_worker_writes_landing_url_to_excel(tmp_path, monkeypatch):
    excel_path = tmp_path / "import.xlsx"
    _make_input_workbook(excel_path)

    monkeypatch.setattr(main_window_module, "MouserClient", _StubMouserClient)
    monkeypatch.setattr(
        main_window_module,
        "download_datasheet_for_part",
        lambda part, hint, client: DownloadResult(
            STATUS_FAILED,
            None,
            "웹 다운로드 실패: HTTP 403",
            "Analog Devices",
            "https://www.analog.com/blocked.pdf",
            "https://www.analog.com/en/products/ad8030.html",
        ),
    )

    rows = [{"row": 2, "part_number": "AD8030ARZ", "manufacturer": None}]
    worker = DatasheetWorker(rows, str(excel_path), "부품리스트")
    worker.run()

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["부품리스트"]
    header_row = [c.value for c in ws[1]]
    col_idx = header_row.index(COL_LANDING_PAGE) + 1
    assert ws.cell(row=2, column=col_idx).value == "https://www.analog.com/en/products/ad8030.html"
    wb.close()
